import csv
from io import BytesIO, StringIO

import openpyxl
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.db import models
from django.http import QueryDict

from bloomerp.models import ApplicationField, FieldPolicy, Policy, RowPolicy, RowPolicyRule
from bloomerp.services.export_services import ExportService
from bloomerp.tests.base import BaseBloomerpModelTestCase
from bloomerp.tests.utils.dynamic_models import create_test_models


class TestExportService(BaseBloomerpModelTestCase):
    def _ensure_permissions_for_model(self, model):
        content_type = ContentType.objects.get_for_model(model)
        for perm in set(model._meta.default_permissions) | {"export"}:
            codename = f"{perm}_{model._meta.model_name}"
            Permission.objects.get_or_create(
                codename=codename,
                content_type=content_type,
                defaults={"name": f"Can {perm} {model._meta.verbose_name}"},
            )

    def test_create_export_bytes_csv_exports_selected_fields_for_admin(self):
        service = ExportService(
            model=self.CustomerModel,
            user=self.admin_user,
            permission_str="export_customer",
        )
        fields = ApplicationField.get_for_model(self.CustomerModel).filter(
            field__in=["first_name", "age"]
        )

        export_bytes, content_type, extension = service.create_export_bytes(
            application_fields=list(fields),
            file_type="csv",
        )

        decoded = export_bytes.decode("utf-8")
        rows = list(csv.reader(StringIO(decoded)))

        self.assertEqual(content_type, "text/csv")
        self.assertEqual(extension, "csv")
        self.assertEqual(rows[0], ["first_name", "age"])
        self.assertEqual(len(rows), self.CustomerModel.objects.count() + 1)

    def test_create_export_bytes_csv_respects_export_row_permissions(self):
        self._ensure_permissions_for_model(self.CustomerModel)

        allowed = self.CustomerModel.objects.create(
            first_name="Allowed",
            last_name="Customer",
            age=77,
        )
        self.CustomerModel.objects.create(
            first_name="Blocked",
            last_name="Customer",
            age=88,
        )

        row_policy = RowPolicy.objects.create(
            content_type=ContentType.objects.get_for_model(self.CustomerModel),
            name="Customer export rows",
        )
        row_rule = RowPolicyRule.objects.create(
            row_policy=row_policy,
            rule={
                "field": "age",
                "application_field_id": ApplicationField.get_by_field(self.CustomerModel, "age").id,
                "operator": "exact",
                "value": "77",
            },
        )
        row_rule.add_permission("export_customer")

        field_policy = FieldPolicy.objects.create(
            content_type=ContentType.objects.get_for_model(self.CustomerModel),
            name="Customer export fields",
            rule={
                str(ApplicationField.get_by_field(self.CustomerModel, "first_name").id): ["export_customer"],
            },
        )
        policy = Policy.objects.create(
            name="Customer export policy",
            row_policy=row_policy,
            field_policy=field_policy,
        )
        policy.assign_user(self.normal_user)

        service = ExportService(
            model=self.CustomerModel,
            user=self.normal_user,
            permission_str="export_customer",
        )
        fields = ApplicationField.get_for_model(self.CustomerModel).filter(field="first_name")

        export_bytes, _, _ = service.create_export_bytes(
            application_fields=list(fields),
            file_type="csv",
        )

        decoded = export_bytes.decode("utf-8")
        rows = list(csv.reader(StringIO(decoded)))

        self.assertEqual(rows[0], ["first_name"])
        self.assertEqual(rows[1:], [["Allowed"]])
        self.assertNotIn("Blocked", decoded)

    def test_create_export_bytes_csv_applies_query_params(self):
        self.CustomerModel.objects.create(
            first_name="FilterMatch",
            last_name="Customer",
            age=77,
        )
        self.CustomerModel.objects.create(
            first_name="FilterMiss",
            last_name="Customer",
            age=88,
        )

        service = ExportService(
            model=self.CustomerModel,
            user=self.admin_user,
            permission_str="export_customer",
        )
        fields = ApplicationField.get_for_model(self.CustomerModel).filter(field="first_name")

        export_bytes, _, _ = service.create_export_bytes(
            application_fields=list(fields),
            file_type="csv",
            query_params=QueryDict("age=77&q=Filter"),
        )

        decoded = export_bytes.decode("utf-8")
        rows = list(csv.reader(StringIO(decoded)))

        self.assertEqual(rows[0], ["first_name"])
        self.assertEqual(rows[1:], [["FilterMatch"]])

    def test_create_export_bytes_csv_supports_pk_alias(self):
        service = ExportService(
            model=self.CustomerModel,
            user=self.admin_user,
            permission_str="export_customer",
        )
        fields = ApplicationField.get_for_model(self.CustomerModel).filter(field="pk")

        export_bytes, _, _ = service.create_export_bytes(
            application_fields=list(fields),
            file_type="csv",
        )

        decoded = export_bytes.decode("utf-8")
        rows = list(csv.reader(StringIO(decoded)))

        self.assertEqual(rows[0], ["pk"])
        self.assertEqual(len(rows), self.CustomerModel.objects.count() + 1)
        self.assertTrue(all(row[0] for row in rows[1:]))


class TestExportServiceRelationships(BaseBloomerpModelTestCase):
    auto_create_customers = False

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.NoteModel = create_test_models(
            app_label="bloomerp",
            model_defs={
                "RelationshipNote": {
                    "customer": models.ForeignKey(cls.CustomerModel, on_delete=models.CASCADE),
                    "name": models.CharField(max_length=100),
                }
            },
            use_bloomerp_base=True,
        )["RelationshipNote"]
        cls._register_dynamic_model_routes([cls.NoteModel])

    def extendedSetup(self):
        self.customer = self.CustomerModel.objects.create(
            first_name="Alice",
            last_name="Example",
            age=31,
        )
        self.note_one = self.NoteModel.objects.create(customer=self.customer, name="Related note 1")
        self.note_two = self.NoteModel.objects.create(customer=self.customer, name="Related note 2")

    def test_create_export_bytes_csv_serializes_one_to_many_ids(self):
        service = ExportService(
            model=self.CustomerModel,
            user=self.admin_user,
            permission_str="export_customer",
        )
        relationship_field = ApplicationField.get_for_model(self.CustomerModel).filter(
            related_model=ContentType.objects.get_for_model(self.NoteModel),
            field_type="OneToManyField",
        ).first()
        self.assertIsNotNone(relationship_field)

        export_bytes, _, _ = service.create_export_bytes(
            application_fields=[relationship_field],
            file_type="csv",
            query_params=QueryDict(f"pk={self.customer.pk}"),
        )

        decoded = export_bytes.decode("utf-8")
        rows = list(csv.reader(StringIO(decoded)))

        self.assertEqual(rows[0], [relationship_field.field])
        self.assertEqual(rows[1][0], f"{self.note_one.pk};{self.note_two.pk}")


class TestExportServiceExcel(BaseBloomerpModelTestCase):
    create_foreign_models = True

    def test_create_export_bytes_xlsx_serializes_uuid_backed_foreign_keys(self):
        customer = self.CustomerModel.objects.first()
        customer.country = self.CountryModel.objects.first()
        customer.save(update_fields=["country"])

        service = ExportService(
            model=self.CustomerModel,
            user=self.admin_user,
            permission_str="export_customer",
        )
        fields = ApplicationField.get_for_model(self.CustomerModel).filter(field="country")

        export_bytes, content_type, extension = service.create_export_bytes(
            application_fields=list(fields),
            file_type="xlsx",
        )

        workbook = openpyxl.load_workbook(BytesIO(export_bytes))
        sheet = workbook.active

        self.assertEqual(
            content_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(extension, "xlsx")
        self.assertEqual(sheet["A1"].value, "country")
        exported_values = [
            "" if cell_value is None else str(cell_value)
            for cell_value in (sheet.cell(row=row_index, column=1).value for row_index in range(2, sheet.max_row + 1))
        ]
        self.assertIn(str(customer.country_id), exported_values)
